# -*- coding: utf-8 -*-
import os
#获取上一级目录
base_path = os.path.abspath(os.path.dirname(os.getcwd()))
# print(base_path)
import sys
sys.path.append(base_path)
import openpyxl

# #拿到所有内容
# open_excel = openpyxl.load_workbook(base_path+'\Case\imooc.xlsx')
# #拿到内容中的sheet页
# sheet_name = open_excel.sheetnames
# #选择其中某一个sheet页
# excel_value = open_excel[sheet_name[0]]
# print(excel_value)
# print(excel_value.cell(1,3).value)
# print(excel_value.max_row)


class HandleExcel():
    """
    读取excel中case
    """
    def load_excel(self):
        """
        加载excel
        """
        open_excel = openpyxl.load_workbook(base_path + '\Case\imooc.xlsx')
        return open_excel

    def get_sheet_data(self,index=None):
        """
        加载sheet页中数据
        :param index: index为sheet页编号，当为默认值时，赋值0
        :return: 返回sheet页对象
        """
        sheet_name = self.load_excel().sheetnames
        if index ==None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self,row,cols,index=None):
        """
        获取某个单元格内容
        :param row:sheet页中的行
        :param cols:sheet页中的列
        :return:但会单元格内容对象
        """
        data = self.get_sheet_data(index).cell(row = row,column = cols).value
        return data
    def get_rows(self):
        """
        获取行数
        :return:
        """
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self,row):
        """
        某一行内容
        :param row:
        :return:
        """
        row_list = []
        for i in  self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list
    def excel_write_data(self,row,cols,value):
        """
        写入数据
        :param row:
        :param cols:
        :param value:
        :return:
        """
        #加载
        wb = self.load_excel()
        #激活
        wr = wb.active
        #写入
        wr.cell(row,cols,value)
        #保存
        wb.save(base_path + '\Case\imooc.xlsx')


    def get_columns_value(self,key=None):
        """
        获取某一列
        :param key: 列号
        :return:某一列值的列表
        """
        if key ==None:
            key = 'A'
        column_list = []
        column_list_data = self.get_sheet_data()[key]
        for i in column_list_data:
            column_list.append(i.value)
        return column_list

    def get_rows_number(self,case_id):
        """
        获取行号
        :return:行号
        """
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id ==col_data:
                print(col_data)
                return num
            num = num+1
        return num



excel_data = HandleExcel()
if __name__ == '__main__':
    handle_excel = HandleExcel()
    #print(handle_excel.get_rows_value(2))
    #print(handle_excel.get_rows())
    #data = {"aaa":"bbb"}
    #handle_excel.excel_write_data(2,13,"成功")
    print(handle_excel.get_columns_value("A"))
    print(handle_excel.get_rows_number('imooc_003'))
    #print(handle_excel.get_rows_value(4))
